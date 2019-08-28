"""
Программа анализирует файлы с результатами экзаменов ОГЭ и ЕГЭ
Сводные данные будут отражены на новой странице в файле
Версия 0.6
"""


from openpyxl import load_workbook as lw
from datetime import datetime
from gui import list_of_files, directory
import os

from settings import *


class ResultFile:
    """
    Файл с результатами экзаменов
    """
    def __init__(self, filename):
        """
        Открывает файл Excel с дополнительными атрибутами:
        name - название файла, который открыт
        sheet - активная страница для работы
        columns - кортеж с номерами столбцов, отвечающих за соответствующие поля
        begin_row - номер строки с первым участником
        subject_cell - ячейка с названием предмета
        subject - название предмета
        students - словарь с данными всех учеников в файле:
            ключ - номер и буква класса (например, 11Н)
            значение - список списков. Один вложенный список - данные одного ученика
            вложенный список - Класс, Ф, И, О, Краткий ответ, Развернутый ответ, Первичный балл, Оценка
        :param filename: путь к файлу Excel. Тип данных - str
        """
        # открываем файл excel
        self.file = lw(filename)
        # запоминаем имя файла
        self.name = filename
        # выбираем активную страницу для работы
        self.sheet = self.file.active
        # проверяем версию файла - для 9 или 11 класса
        page_version = self.set_page_version()
        # выбираем нужное из словаря с параметрами
        self.columns = versions[page_version]['columns']  # номера столбцов
        self.begin_row = versions[page_version]['begin_row']  # первая строка с данными ученика
        self.subject_cell = versions[page_version]['subject_cell']  # ячейка с названием предмета
        # запоминаем название предмета
        self.subject = self.get_subject()
        # набираем словарь с данными учеников
        self.students = self.get_all_students()

    def set_page_version(self):
        """
        Возвращает версию файла в виде ключа для словаря versions
        :return: 'nine', 'eleven' или 'appeal' в зависимости от версии файла
        """
        check = self.sheet[cell_check_version].value
        if check == 13273:
            return 'nine'
        elif check is None:
            return 'eleven'
        else:
            return 'appeal'

    def get_subject(self):
        """
        Вытаскивает название предмета из ячейки subject_cell
        :return: Название предмета. Тип данных - str
        """
        subject = ''
        cell = [i for i in self.sheet.cell[self.subject_cell]]
        if len[cell[0]] > 2:
            subject = str(cell[0]) + str(cell[1])
        else:
            cell = cell[2:]
            cell = cell[:-1]
            for i in cell:
                subject += str(i)
        return subject

    def get_student(self, row):
        """
        Набирает в список данные одного ученика
        :param row: номер строки, на которой находятся данные ученика. Тип данных - int
        :return: список в порядке Класс, Ф, И, О, Краткий ответ, Развернутый ответ, Первичный балл, Оценка
        """
        student = []
        for i in self.columns:
            student.append(self.sheet.cell(row=row, column=i).value)
        return student

    def write_student_to_dict(self, one_student, students):
        """
        Добавляет данные одного ученика в словарь.
        Если ключа с номером класса нет - создает его.
        :param one_student: список с данными одного ученика
        :param students: словарь с ключами в виде классов (например, 11Н)
        :return: словарь students с новым значением
        """
        grade = one_student[0]    # Номер класса (ключ словаря)
        student_data = one_student[1:]    # Набор данных ученика

        # Если класса еще нет в словаре - добавляем ключ
        if grade not in students.keys():
            students[grade] = []

        # обновляем список по ключу grade новым набором данных
        students[grade].append(student_data)
        return students

    def get_all_students(self):
        """
        Набирает всех учеников из файла в словарь
        :return: словарь:
        Ключ - номер класса (например, 11Н)
        Значения - списки с данными по каждому сдававшему ученику из класса
        """
        students = {}
        for row in range(self.begin_row, self.sheet.max_row + 1):
            one_student = self.get_student(row)
            students = self.write_student_to_dict(one_student, students)
        # Удаляем из словаря пустые включения
        students.pop(None, None)
        return students


class ResultTable:
    """
    Сводная таблица с результатами по предметам
    """
    def __init__(self):
        pass


def analise_results():
    for filename in list_of_files:
        file = ResultFile(filename)
        print(file.name, file.set_page_version())


if __name__ == '__main__':
    analise_results()
