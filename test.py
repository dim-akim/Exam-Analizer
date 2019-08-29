"""
Модуль для тестирования
"""

from main import *
from openpyxl import load_workbook as lw


def test1():
    # Проверка работы параметра max_column
    file = lw('Результаты 2019/Пример.xlsx')
    page = file.active

    for i in range(1, page.max_column + 1):
        print(page.cell(row=11,  column=i).value)


def test2():
    # Проверяет, работает ли класс ResultFile
    filename = 'Результаты 2019/Пример.xlsx'
    file = ResultFile(filename)
    # print(file.name)
    # print(file.sheet)
    count = 0
    for i in file.students.keys():
        print(i)
        for j in file.students[i]:
            count += 1
            print(j)
    print(count)


def test3():
    # Проверяет работу ячейки-маркера типа файла
    filenames = [
        'Пример_11.xlsx',
        'Пример_аппел.xlsx',
        'Пример.xlsx'
    ]
    cell = 'B10'
    for name in filenames:
        book = lw(name)
        page = book.active
        print(page[cell].value)


def test4():
    # Проверяет вытаскивание названия предмета
    filenames = [
        'Пример_11.xlsx',
        'Пример_аппел.xlsx',
        'Пример.xlsx'
    ]
    for name in filenames:
        book = ResultFile(name)
        print(book.subject)


def test5():
    # Проверяет новый подход в считывании таблицы и определении версии
    filenames = [
        'Пример_11.xlsx',
        'Пример_аппел.xlsx',
        'Пример.xlsx'
    ]
    for name in filenames:
        book = lw(name)
        page = book.active
        for row in range(1, page.max_row):
            a = page.cell(row=row, column=1).value
            # print(row, a)
            if a == '№':
                row += 1
                print(row)
                break
        student = []
        for i in range(1, page.max_column + 1):
            a = page.cell(row=row, column=i).value
            if a is not None:
                student.append(a)
        print(student)

# test1()
# test2()
# test3()
test5()
